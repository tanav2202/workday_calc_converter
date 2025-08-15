import gradio as gr
import pandas as pd
import re
from datetime import datetime, timedelta
from icalendar import Calendar, Event
import uuid
import pytz
import openpyxl
from io import BytesIO
import tempfile
import os

def excel_date_to_datetime(excel_date):
    """Convert Excel serial date to Python datetime."""
    if isinstance(excel_date, (int, float)):
        # Excel's epoch starts from 1900-01-01 (with 1900 incorrectly treated as leap year)
        return datetime(1899, 12, 30) + timedelta(days=excel_date)
    return excel_date

def parse_meeting_pattern(meeting_pattern):
    """Parse the meeting pattern string to extract schedule information."""
    if not meeting_pattern or pd.isna(meeting_pattern):
        return []
    
    # Split multiple patterns (some courses have multiple time blocks)
    patterns = meeting_pattern.split('\n\n')
    schedules = []
    
    for pattern in patterns:
        pattern = pattern.strip()
        if not pattern:
            continue
            
        # Parse pattern like: "2025-11-17 - 2025-12-17 | Mon Wed | 9:30 a.m. - 11:00 a.m. | UBCV | Hugh Dempster Pavilion (DMP) | Floor: 1 | Room: 110"
        parts = pattern.split(' | ')
        
        if len(parts) >= 6:
            date_range = parts[0].strip()
            days = parts[1].strip()
            time_range = parts[2].strip()
            campus = parts[3].strip()
            building = parts[4].strip()
            location_parts = parts[5:]  # Everything from floor onwards
            
            # Parse date range
            try:
                start_date_str, end_date_str = date_range.split(' - ')
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except Exception as e:
                continue
            
            # Parse time range
            try:
                time_parts = time_range.split(' - ')
                start_time_str = time_parts[0].strip()
                end_time_str = time_parts[1].strip()
                
                # Normalize time format - handle both "p.m." and "PM" formats
                def normalize_time_string(time_str):
                    time_str = time_str.replace('a.m.', 'AM').replace('p.m.', 'PM')
                    time_str = time_str.replace('a.m', 'AM').replace('p.m', 'PM')
                    return time_str
                
                start_time_str = normalize_time_string(start_time_str)
                end_time_str = normalize_time_string(end_time_str)
                
                # Convert 12-hour format to 24-hour format
                start_time = datetime.strptime(start_time_str, '%I:%M %p').time()
                end_time = datetime.strptime(end_time_str, '%I:%M %p').time()
            except Exception as e:
                continue
            
            # Parse days
            day_mapping = {
                'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6
            }
            
            weekdays = []
            for day_name in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']:
                if day_name in days:
                    weekdays.append(day_mapping[day_name])
            
            # Create location string
            location = f"{building}, {' | '.join(location_parts)}, {campus}".strip(', ')
            
            if weekdays:  # Only add if we have valid weekdays
                schedules.append({
                    'start_date': start_date,
                    'end_date': end_date,
                    'start_time': start_time,
                    'end_time': end_time,
                    'weekdays': weekdays,
                    'location': location
                })
    
    return schedules

def read_course_data_from_excel(excel_file_path):
    """Read course data from Excel file using openpyxl for better cell access."""
    
    # Load workbook
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook['View Courses for Student']
    
    # Extract headers from row 3
    headers = []
    for col in range(1, 15):  # Columns A to N
        cell_value = worksheet.cell(row=3, column=col).value
        headers.append(cell_value if cell_value else '')
    
    # Extract course data from rows 4 onwards
    courses = []
    for row_num in range(4, 100):  # Check more rows for web version
        row_data = []
        has_data = False
        
        for col in range(1, len(headers) + 1):
            cell_value = worksheet.cell(row=row_num, column=col).value
            if cell_value is not None:
                row_data.append(cell_value)
                has_data = True
            else:
                row_data.append('')
        
        # If row has course data (course listing in column B)
        if has_data and len(row_data) > 1 and row_data[1]:
            courses.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(courses, columns=headers)
    
    # Clean up column names
    df.columns = [col.strip() if col else f'Column_{i}' for i, col in enumerate(df.columns)]
    
    return df

def create_ics_from_excel_df(df):
    """Convert DataFrame to ICS format using icalendar library."""
    
    if df.empty:
        return None, "No course data found in the Excel file!"
    
    # Create calendar
    cal = Calendar()
    cal.add('prodid', '-//Course Schedule//Course Schedule//EN')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('method', 'PUBLISH')
    cal.add('x-wr-calname', 'UBC Course Schedule')
    cal.add('x-wr-timezone', 'America/Vancouver')
    
    # Set timezone
    vancouver_tz = pytz.timezone('America/Vancouver')
    
    events_created = 0
    
    # Process each course
    for index, row in df.iterrows():
        course_listing = str(row.get('Course Listing', '')).strip()
        meeting_patterns = str(row.get('Meeting Patterns', '')).strip()
        
        if not course_listing or not meeting_patterns or meeting_patterns == 'nan':
            continue
            
        section = str(row.get('Section', '')).strip()
        instructor = str(row.get('Instructor', '')).strip()
        instructional_format = str(row.get('Instructional Format', '')).strip()
        
        # Parse meeting patterns
        schedules = parse_meeting_pattern(meeting_patterns)
        
        if not schedules:
            continue
        
        for schedule in schedules:
            # Generate events for each occurrence
            current_date = schedule['start_date']
            while current_date <= schedule['end_date']:
                if current_date.weekday() in schedule['weekdays']:
                    # Create event
                    event = Event()
                    
                    # Combine date and time
                    start_datetime = datetime.combine(current_date, schedule['start_time'])
                    end_datetime = datetime.combine(current_date, schedule['end_time'])
                    
                    # Localize to Vancouver timezone
                    start_datetime = vancouver_tz.localize(start_datetime)
                    end_datetime = vancouver_tz.localize(end_datetime)
                    
                    # Set event properties
                    event.add('uid', str(uuid.uuid4()))
                    event.add('dtstart', start_datetime)
                    event.add('dtend', end_datetime)
                    event.add('dtstamp', datetime.now(vancouver_tz))
                    
                    # Create summary
                    summary = course_listing
                    if instructional_format:
                        summary += f" - {instructional_format}"
                    event.add('summary', summary)
                    
                    # Create description
                    description_parts = []
                    if section:
                        description_parts.append(f"Section: {section}")
                    if instructor:
                        description_parts.append(f"Instructor: {instructor}")
                    if instructional_format:
                        description_parts.append(f"Format: {instructional_format}")
                    
                    if description_parts:
                        event.add('description', '\\n'.join(description_parts))
                    
                    # Set location
                    event.add('location', schedule['location'])
                    
                    # Add categories
                    event.add('categories', ['EDUCATION', 'COURSE'])
                    
                    # Add to calendar
                    cal.add_component(event)
                    events_created += 1
                
                current_date += timedelta(days=1)
    
    return cal.to_ical(), events_created

def process_excel_file(file):
    """Main function to process uploaded Excel file."""
    if file is None:
        return None, "❌ Please upload an Excel file."
    
    try:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(file)
            tmp_file_path = tmp_file.name
        
        # Read course data
        df = read_course_data_from_excel(tmp_file_path)
        
        if df.empty:
            return None, "❌ No course data found in the Excel file. Please check the file format."
        
        # Create iCal content using icalendar library
        ical_content, events_created = create_ics_from_excel_df(df)
        
        if ical_content is None:
            return None, "❌ Failed to create calendar events."
        
        # Save iCal file
        output_path = tempfile.NamedTemporaryFile(delete=False, suffix='.ics', mode='wb')
        output_path.write(ical_content)
        output_path.close()
        
        # Clean up temporary Excel file
        os.unlink(tmp_file_path)
        
        success_message = f"""✅ **Success!** 

📊 **Results:**
- Found **{len(df)}** course entries
- Created **{events_created}** calendar events
- Generated ICS file ready for download

📱 **Compatible with:**
- Google Calendar
- Outlook
- Apple Calendar
- Any calendar app that supports ICS files

💡 **Next steps:** Download the file and import it into your calendar app!"""
        
        return output_path.name, success_message
        
    except Exception as e:
        return None, f"❌ **Error processing file:** {str(e)}"

# Create Gradio interface
def create_app():
  
    
    with gr.Blocks(
        title="Excel to ICS Calendar Converter",
        theme=gr.themes.Default()
    ) as demo:
        
        with gr.Column(elem_classes="main-content"):
            gr.HTML("""
            <div class="main-header">
                <h1>📅 Excel to ICS Calendar Converter</h1>
                <p>Transform your course schedule into a calendar file instantly</p>
            </div>
            """)
            
            with gr.Row():
                with gr.Column(scale=1):
                    gr.HTML("""
                    <div class="feature-box">
                        <h3>🚀 Features</h3>
                        <ul>
                            <li>📁 Upload Excel course schedules</li>
                            <li>🔄 Automatic pattern parsing</li>
                            <li>📅 Generate ICS calendar files</li>
                            <li>📱 Works with all major calendar apps</li>
                            <li>⏰ Handles multiple time blocks</li>
                            <li>📍 Includes locations & instructors</li>
                        </ul>
                    </div>
                    """)
                    
                    gr.HTML("""
                    <div class="feature-box">
                        <h3>📋 How to Use</h3>
                        <ol>
                            <li>📤 Upload your Excel course file</li>
                            <li>🔄 Click "Convert to Calendar"</li>
                            <li>⬇️ Download the ICS file</li>
                            <li>📱 Import into your calendar app</li>
                        </ol>
                    </div>
                    """)
                
                with gr.Column(scale=2):
                    file_input = gr.File(
                        label="📁 Upload Excel Course Schedule",
                        file_types=[".xlsx", ".xls"],
                        type="binary"
                    )
                    
                    convert_btn = gr.Button(
                        "🔄 Convert to Calendar",
                        variant="primary",
                        size="lg"
                    )
                    
                    status_output = gr.Markdown(
                        label="📊 Status",
                        value="🔼 Upload an Excel file to get started.",
                        elem_classes="status-output"
                    )
                    
                    download_output = gr.File(
                        label="⬇️ Download Calendar File",
                        visible=False
                    )
            
            def process_and_update(file):
                ics_file, message = process_excel_file(file)
                if ics_file:
                    return message, gr.File(value=ics_file, visible=True)
                else:
                    return message, gr.File(visible=False)
            
            convert_btn.click(
                fn=process_and_update,
                inputs=[file_input],
                outputs=[status_output, download_output]
            )
            
            gr.HTML("""
            <div class="footer">
                <p>🛠️ 🏫 Optimized for UBC schedules on workday </p>
                <p>📧 Having issues? Check that your Excel file has the course data in the expected format (directly downloaded from workday).</p>
            </div>
            """)
    
    return demo

# Entry point for Vercel
app = create_app()

if __name__ == "__main__":
    app.launch()